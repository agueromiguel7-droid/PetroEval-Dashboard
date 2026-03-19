import openpyxl

file_path = "MEIPO - SOLUCIÓN.xlsm"
output_file = "formulas.txt"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb["Eval. Econ. Proy"]
    
    with open(output_file, "w", encoding="utf-8") as f:
        # Extract variables from top rows (e.g., rows 1-29)
        f.write("--- VARIABLES (Rows 1-29) ---\n")
        for row in range(1, 30):
            values = []
            for col in range(1, 40):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    values.append(f"{cell.coordinate}: {cell.value}")
            if values:
                f.write(" | ".join(values) + "\n")
                
        # Extract formulas from rows 30 to 80
        f.write("\n--- CALCULATIONS (Rows 30-80, Cols 8-40) ---\n")
        for row in range(30, 81):
            formulas = []
            for col in range(8, 40):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    formulas.append(f"{cell.coordinate}: {cell.value}")
            if formulas:
                f.write(" | ".join(formulas) + "\n")
                
        f.write("\nSuccess! Saved formulas.\n")
        print("Done.")

except Exception as e:
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(f"Error: {str(e)}")
    print(f"Error: {e}")
